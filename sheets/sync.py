"""Синхронизация с Google Sheets и Excel"""

import os
import json
import logging
import asyncio
from datetime import datetime
from google.auth.transport.requests import Request
from google.oauth2.service_account import Credentials
from googleapiclient import discovery
from config import GOOGLE_SHEETS_CREDENTIALS_JSON, GOOGLE_SHEETS_ID
from database.supabase_client import db

logger = logging.getLogger(__name__)

# Scopes для Google Sheets
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']


def _build_sheets_service():
    """Синхронная функция создания сервиса Google Sheets"""
    if not os.path.exists(GOOGLE_SHEETS_CREDENTIALS_JSON):
        raise FileNotFoundError(f"❌ Файл credentials не найден: {GOOGLE_SHEETS_CREDENTIALS_JSON}")
    
    credentials = Credentials.from_service_account_file(
        GOOGLE_SHEETS_CREDENTIALS_JSON,
        scopes=SCOPES
    )
    
    service = discovery.build('sheets', 'v4', credentials=credentials)
    return service


async def get_google_sheets_service():
    """Получить сервис Google Sheets (неблокирующий вызов)"""
    return await asyncio.to_thread(_build_sheets_service)


async def sync_purchases_to_sheets(user_id: int):
    """Синхронизировать покупки на Google Sheets (неблокирующий вызов)"""
    try:
        logger.info(f"🔄 Начинаем синхронизацию для user_id={user_id}")
        
        # Получить все покупки
        purchases = await db.get_all_purchases(user_id)
        drinks = await db.get_all_drinks(user_id)
        
        logger.info(f"📊 Получено {len(purchases)} покупок и {len(drinks)} напитков")
        
        # Подготовить данные
        purchase_rows = [["Дата покупки", "Тип напитка", "Объем (мл)", "Цена (₽)", 
                         "Калории (итого)", "Сахар (г итого)", "Кофеин (мг итого)", 
                         "Натрий (мг итого)", "Примечание"]]
        
        for p in purchases:
            drink_name = p['drinks']['name'] if isinstance(p.get('drinks'), dict) else 'Неизвестно'
            row = [
                p['purchase_date'],
                drink_name,
                p['volume_ml'],
                p.get('price_rub', 0),
                p.get('calories_total', 0),
                p.get('sugar_total', 0),
                p.get('caffeine_total', 0),
                p.get('sodium_total', 0),
                p.get('notes', '')
            ]
            purchase_rows.append(row)
        
        # Подготовить справочник напитков
        drink_rows = [["Напиток", "На 100мл - Калории", "На 100мл - Сахар (г)", 
                      "На 100мл - Кофеин (мг)", "На 100мл - Натрий (мг)"]]
        
        for d in drinks:
            row = [
                d['name'],
                d['calories_per_100ml'],
                d['sugar_per_100ml'],
                d['caffeine_per_100ml'],
                d['sodium_per_100ml']
            ]
            drink_rows.append(row)
        
        logger.info(f"📝 Подготовлено {len(purchase_rows)-1} строк покупок")
        logger.info(f"📝 Подготовлено {len(drink_rows)-1} строк напитков")
        
        # Получить сервис
        service = await get_google_sheets_service()
        
        if not GOOGLE_SHEETS_ID:
            logger.error("❌ GOOGLE_SHEETS_ID не установлен")
            raise ValueError("GOOGLE_SHEETS_ID должен быть в .env")
        
        # Обновить лист покупок
        requests = [
            {
                "updateSheet": {
                    "sheetProperties": {
                        "sheetId": 0,  # Первый лист
                        "title": "Покупки"
                    }
                }
            },
            {
                "deleteRange": {
                    "range": {
                        "sheetId": 0,
                        "startRowIndex": 1,
                        "endRowIndex": 1000
                    },
                    "shiftDimension": "ROWS"
                }
            },
            {
                "appendCells": {
                    "sheetId": 0,
                    "rows": [{"values": [{"userEnteredValue": {"stringValue": str(cell)}} 
                                        for cell in row]} for row in purchase_rows],
                    "fields": "userEnteredValue"
                }
            }
        ]
        
        # Отправить запрос
        body = {"requests": requests}
        
        logger.info("📤 Отправляем данные на Google Sheets...")
        
        # Запускаем в отдельном потоке, чтобы не блокировать event loop
        await asyncio.to_thread(
            service.spreadsheets().batchUpdate(
                spreadsheetId=GOOGLE_SHEETS_ID,
                body=body
            ).execute
        )
        
        logger.info(f"✅ Синхронизация завершена успешно")
        await db.log_sync(user_id, 'success', f'Синхронизировано {len(purchases)} покупок')
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Ошибка синхронизации: {e}")
        await db.log_sync(user_id, 'failed', str(e))
        raise


def _write_excel_sync(purchases, drinks, filepath):
    """Синхронная функция создания и сохранения Excel файла с сохранением форматирования"""
    import openpyxl
    
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "Покупки"
    
    # Получить или создать листы
    if "Покупки" in wb.sheetnames:
        ws_p = wb["Покупки"]
    else:
        ws_p = wb.create_sheet("Покупки")
        
    if "Состав напитков" in wb.sheetnames:
        ws_d = wb["Состав напитков"]
    else:
        ws_d = wb.create_sheet("Состав напитков")
        
    if "Статистика" in wb.sheetnames:
        ws_s = wb["Статистика"]
    else:
        ws_s = wb.create_sheet("Статистика")
    
    # Очистить старые данные (начиная со 2 строки), сохраняя стили
    for r in range(2, 2000):
        any_val = False
        for c in range(1, 10):
            if ws_p.cell(row=r, column=c).value is not None:
                ws_p.cell(row=r, column=c).value = None
                any_val = True
        if not any_val and r > 105:
            break
            
    for r in range(2, 500):
        any_val = False
        for c in range(1, 6):
            if ws_d.cell(row=r, column=c).value is not None:
                ws_d.cell(row=r, column=c).value = None
                any_val = True
        if not any_val and r > 55:
            break
            
    # Записать покупки
    for idx, p in enumerate(purchases):
        row_idx = idx + 2
        drink_name = p['drinks']['name'] if isinstance(p.get('drinks'), dict) else 'Неизвестно'
        ws_p.cell(row=row_idx, column=1, value=p['purchase_date'])
        ws_p.cell(row=row_idx, column=2, value=drink_name)
        ws_p.cell(row=row_idx, column=3, value=p['volume_ml'])
        ws_p.cell(row=row_idx, column=4, value=p.get('price_rub', 0))
        ws_p.cell(row=row_idx, column=5, value=p.get('calories_total', 0))
        ws_p.cell(row=row_idx, column=6, value=p.get('sugar_total', 0))
        ws_p.cell(row=row_idx, column=7, value=p.get('caffeine_total', 0))
        ws_p.cell(row=row_idx, column=8, value=p.get('sodium_total', 0))
        ws_p.cell(row=row_idx, column=9, value=p.get('notes', ''))
        
    # Записать состав напитков
    for idx, d in enumerate(drinks):
        row_idx = idx + 2
        ws_d.cell(row=row_idx, column=1, value=d['name'])
        ws_d.cell(row=row_idx, column=2, value=d['calories_per_100ml'])
        ws_d.cell(row=row_idx, column=3, value=d['sugar_per_100ml'])
        ws_d.cell(row=row_idx, column=4, value=d['caffeine_per_100ml'])
        ws_d.cell(row=row_idx, column=5, value=d['sodium_per_100ml'])
        
    # Рассчитать и записать статистику
    total_purchases = len(purchases)
    total_spent = sum(p.get('price_rub', 0) for p in purchases if p.get('price_rub'))
    avg_price = total_spent / total_purchases if total_purchases > 0 else 0
    total_volume_l = sum(p['volume_ml'] for p in purchases) / 1000
    avg_volume = sum(p['volume_ml'] for p in purchases) / total_purchases if total_purchases > 0 else 0
    total_calories = sum(p.get('calories_total', 0) for p in purchases if p.get('calories_total'))
    total_sugar = sum(p.get('sugar_total', 0) for p in purchases if p.get('sugar_total'))
    total_caffeine = sum(p.get('caffeine_total', 0) for p in purchases if p.get('caffeine_total'))
    total_sodium = sum(p.get('sodium_total', 0) for p in purchases if p.get('sodium_total'))
    
    ws_s.cell(row=1, column=1, value="СТАТИСТИКА ПОКУПОК")
    ws_s.cell(row=2, column=1, value="Всего покупок:")
    ws_s.cell(row=2, column=2, value=total_purchases)
    ws_s.cell(row=3, column=1, value="Всего потрачено (₽):")
    ws_s.cell(row=3, column=2, value=total_spent)
    ws_s.cell(row=4, column=1, value="Средняя цена за покупку (₽):")
    ws_s.cell(row=4, column=2, value=round(avg_price, 2))
    ws_s.cell(row=5, column=1, value="Общее количество выпито (л):")
    ws_s.cell(row=5, column=2, value=round(total_volume_l, 2))
    ws_s.cell(row=6, column=1, value="Средний объем на покупку (мл):")
    ws_s.cell(row=6, column=2, value=round(avg_volume, 1))
    ws_s.cell(row=7, column=1, value="Общие калории (во всех покупках):")
    ws_s.cell(row=7, column=2, value=round(total_calories, 1))
    ws_s.cell(row=8, column=1, value="Общий сахар (г):")
    ws_s.cell(row=8, column=2, value=round(total_sugar, 1))
    ws_s.cell(row=9, column=1, value="Общий кофеин (мг):")
    ws_s.cell(row=9, column=2, value=round(total_caffeine, 1))
    ws_s.cell(row=10, column=1, value="Общий натрий (мг):")
    ws_s.cell(row=10, column=2, value=round(total_sodium, 1))
    
    # Сохранить файл
    wb.save(filepath)


async def export_to_excel(user_id: int, filepath: str):
    """Экспортировать покупки в Excel (неблокирующий вызов)"""
    try:
        logger.info(f"📊 Начинаем экспорт в Excel для user_id={user_id} в путь {filepath}")
        
        # Получить данные
        purchases = await db.get_all_purchases(user_id)
        drinks = await db.get_all_drinks(user_id)
        
        # Выполнить сохранение в фоновом потоке
        await asyncio.to_thread(_write_excel_sync, purchases, drinks, filepath)
        
        logger.info(f"✅ Экспорт завершён: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"❌ Ошибка экспорта: {e}")
        raise
